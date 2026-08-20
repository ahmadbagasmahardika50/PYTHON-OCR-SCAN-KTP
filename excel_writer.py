# -*- coding: utf-8 -*-
"""
excel_writer.py
Menyusun hasil batch processing menjadi file Excel (.xlsx) yang rapi,
memakai openpyxl langsung (styling header, lebar kolom, freeze panes,
highlight baris yang perlu diverifikasi).
"""

import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

logger = logging.getLogger("ktp_ocr.excel_writer")

COLUMNS = [
    ("No", 4),
    ("Nama File Original", 26),
    ("Path Foto Enhanced", 30),
    ("NIK", 18),
    ("Nama", 22),
    ("Tempat/Tgl Lahir", 20),
    ("Jenis Kelamin", 12),
    ("Gol. Darah", 10),
    ("Alamat", 28),
    ("RT/RW", 10),
    ("Kel/Desa", 16),
    ("Kecamatan", 16),
    ("Agama", 10),
    ("Status Perkawinan", 16),
    ("Pekerjaan", 16),
    ("Kewarganegaraan", 14),
    ("Berlaku Hingga", 14),
    ("Provinsi", 14),
    ("Kabupaten/Kota", 16),
    ("Tipe Gambar", 12),
    ("Kartu Terdeteksi (Perspective)", 14),
    ("Metode Super Resolution", 16),
    ("Sudut Koreksi (deskew)", 12),
    ("Skor Kelengkapan (%)", 12),
    ("Status Scan", 16),
    ("Catatan", 34),
    ("Waktu Proses (detik)", 12),
]


def build_row(idx, record):
    """
    record: dict hasil dari main.py (gabungan output preprocessing, SR, OCR,
    parser, plus metadata file). Field yang tidak ada akan diisi kosong.
    """
    return [
        idx,
        record.get("original_filename", ""),
        record.get("enhanced_path", ""),
        record.get("NIK", ""),
        record.get("Nama", ""),
        record.get("Tempat_Tgl_Lahir", ""),
        record.get("Jenis_Kelamin", ""),
        record.get("Gol_Darah", ""),
        record.get("Alamat", ""),
        record.get("RT_RW", ""),
        record.get("Kel_Desa", ""),
        record.get("Kecamatan", ""),
        record.get("Agama", ""),
        record.get("Status_Perkawinan", ""),
        record.get("Pekerjaan", ""),
        record.get("Kewarganegaraan", ""),
        record.get("Berlaku_Hingga", ""),
        record.get("Provinsi", ""),
        record.get("Kabupaten_Kota", ""),
        record.get("tipe_gambar", ""),
        record.get("kartu_terdeteksi", ""),
        record.get("sr_method", ""),
        record.get("deskew_angle", ""),
        record.get("_fill_ratio_pct", ""),
        record.get("_status", "Error"),
        record.get("_catatan", ""),
        record.get("elapsed_sec", ""),
    ]


def write_excel(records, output_path):
    """
    records: list of dict (satu dict per file yang diproses, termasuk yang
    error -- baris error tetap ditulis supaya tidak ada file yang "hilang"
    dari laporan).
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Hasil Scan KTP"

    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    review_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    error_fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
    valid_fill = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")

    for col_idx, (name, _) in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    status_col_idx = [i for i, (name, _) in enumerate(COLUMNS, start=1) if name == "Status Scan"][0]

    for row_idx, record in enumerate(records, start=2):
        row_values = build_row(row_idx - 1, record)
        for col_idx, val in enumerate(row_values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.alignment = Alignment(vertical="top", wrap_text=True)

        status = ws.cell(row=row_idx, column=status_col_idx).value
        if status == "Error":
            fill = error_fill
        elif status == "Perlu Verifikasi":
            fill = review_fill
        elif status == "Valid":
            fill = valid_fill
        else:
            fill = None
        if fill:
            for col_idx in range(1, len(COLUMNS) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = fill

    for i, (_, width) in enumerate(COLUMNS, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    logger.info("File Excel disimpan: %s (%d baris)", output_path, len(records))
